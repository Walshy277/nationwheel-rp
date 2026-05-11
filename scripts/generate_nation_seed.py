import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORLD_DB = ROOT / "-NW- World Database (2).xlsx"
S2_BASE = ROOT / "S2 Base (3).xlsx"
OUT = ROOT / "supabase-nation-seed.sql"


def slugify(value):
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def plain(value):
    text = clean(value)
    if not text:
        return ""
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" -/")
    return "" if re.fullmatch(r"\?+", text) else text


def parse_number_text(value):
    text = clean(value).replace(",", ".")
    if not text:
        return None
    match = re.search(r"([\d.]+)\s*([kmbt])?", text, re.I)
    if not match:
        digits = re.sub(r"[^\d]", "", text)
        return int(digits) if digits else None
    number = float(match.group(1))
    suffix = (match.group(2) or "").lower()
    mult = {"k": 1_000, "m": 1_000_000, "b": 1_000_000_000, "t": 1_000_000_000_000}.get(suffix, 1)
    return int(number * mult)


def parse_spaced_number(value):
    text = clean(value)
    if not text:
        return None
    match = re.search(r"([\d\s]+)", text)
    if not match:
        return parse_number_text(text)
    digits = re.sub(r"\s+", "", match.group(1))
    return int(digits) if digits else None


def parse_gdp(value):
    text = clean(value).replace(",", ".")
    if not text:
        return None
    match = re.search(r"([\d.]+)\s*([bmt])", text, re.I)
    if not match:
        return None
    number = float(match.group(1))
    mult = {"b": 1_000_000_000, "m": 1_000_000, "t": 1_000_000_000_000}[match.group(2).lower()]
    return int(number * mult)


def parse_rank(value):
    text = clean(value)
    match = re.search(r"(\d+)\s*/\s*(?:10|11)", text)
    return int(match.group(1)) if match else None


def parse_hdi(value):
    text = clean(value).replace(",", ".")
    match = re.search(r"0?\.\d+", text)
    return float(match.group(0)) if match else None


def sql(value):
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def compact_lines(parts):
    return "\n".join(f"{k}: {v}" for k, v in parts if clean(v))


def world_database_rows():
    wb = load_workbook(WORLD_DB, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    nations = []
    for row in rows:
        if not any(clean(c) for c in row):
            continue
        spin, name, pop, government, area, economy, military, status, bloc, phase2 = row[:10]
        name = clean(name)
        if not name:
            continue
        bio = compact_lines([
            ("Source", "Season 1 world database"),
            ("Spin", plain(spin)),
            ("Military", plain(military)),
            ("Phase 2", plain(phase2)),
        ])
        nations.append({
            "name": name,
            "slug": slugify(name),
            "government": plain(government) or None,
            "population": parse_number_text(pop),
            "land_km2": parse_number_text(area),
            "gdp_usd": parse_gdp(economy),
            "army_rank": parse_rank(military),
            "economy": plain(economy) or None,
            "diplomatic_status": plain(status) or None,
            "bloc": plain(bloc) or None,
            "hdi": None,
            "bio": bio,
        })
    return nations


def s2_base_rows():
    wb = load_workbook(S2_BASE, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    nations = []
    for row in rows:
        if not any(clean(c) for c in row):
            continue
        spin, name, size, population, government, traits, economy, gdp, hdi, military, cultures, allegiance, blessing, loc, influence, note = row[:16]
        name = clean(name)
        if not name:
            continue
        bio = compact_lines([
            ("Source", "Season 2 base"),
            ("Spin", plain(spin)),
            ("Leader traits", plain(traits)),
            ("Military", plain(military)),
            ("Cultures", plain(cultures)),
            ("Allegiance", plain(allegiance)),
            ("Blessing / Curse", plain(blessing)),
            ("Localisation", plain(loc)),
            ("Influence", plain(influence)),
            ("Note", plain(note)),
        ])
        nations.append({
            "name": name,
            "slug": slugify(name),
            "government": plain(government) or None,
            "population": parse_spaced_number(population),
            "land_km2": parse_spaced_number(size),
            "gdp_usd": parse_gdp(gdp),
            "army_rank": parse_rank(military),
            "economy": plain(economy) or None,
            "diplomatic_status": plain(blessing) or None,
            "bloc": plain(allegiance) or None,
            "hdi": parse_hdi(hdi),
            "bio": bio,
        })
    return nations


def main():
    nations = world_database_rows() + s2_base_rows()
    seen = {}
    for nation in nations:
        # Later source rows win on duplicate slugs, with bios combined.
        if nation["slug"] in seen:
            previous = seen[nation["slug"]]
            nation["bio"] = previous["bio"] + "\n\n" + nation["bio"]
        seen[nation["slug"]] = nation
    nations = list(seen.values())

    cols = [
        "name", "slug", "government", "population", "gdp_usd", "land_km2",
        "army_rank", "hdi", "economy", "diplomatic_status", "bloc", "bio",
    ]
    values = []
    for n in nations:
        values.append("(" + ", ".join(sql(n[c]) for c in cols) + ")")

    output = [
        "-- Generated from -NW- World Database (2).xlsx and S2 Base (3).xlsx.",
        "-- Run after the main Supabase schema setup.",
        "",
        "insert into nations (" + ", ".join(cols) + ") values",
        ",\n".join(values),
        "on conflict (slug) do update set",
        "  name = excluded.name,",
        "  government = excluded.government,",
        "  population = excluded.population,",
        "  gdp_usd = excluded.gdp_usd,",
        "  land_km2 = excluded.land_km2,",
        "  army_rank = excluded.army_rank,",
        "  hdi = excluded.hdi,",
        "  economy = excluded.economy,",
        "  diplomatic_status = excluded.diplomatic_status,",
        "  bloc = excluded.bloc,",
        "  bio = excluded.bio;",
        "",
    ]
    OUT.write_text("\n".join(output), encoding="utf-8")
    print(f"Wrote {len(nations)} nations to {OUT.name}")


if __name__ == "__main__":
    main()
